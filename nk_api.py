import os
import sys
import argparse
import logging
import requests
from datetime import datetime
import csv
from openpyxl import load_workbook


# ---------------------------
# Настройка логгера
# ---------------------------
def setup_logging(log_file: str = None):
    """
    Настройка логирования с выводом в консоль и опционально в файл
    """
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Форматтер для логов
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # Обработчик для консоли
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    # Удаляем существующие обработчики
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # Добавляем консольный обработчик
    logger.addHandler(console_handler)
    
    # Если указан файл для логирования, добавляем файловый обработчик
    if log_file:
        # Создаем директорию для логов, если её нет
        log_dir = os.path.dirname(log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
            
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        logger.info(f"Логирование в файл: {log_file}")


logger = logging.getLogger(__name__)


class NK:
    """
    Клиент API Национального Каталога маркированных товаров (Честный Знак).
    Версия API: v5.38
    """

    def __init__(self, token: str = None, apikey: str = None, sandbox: bool = False):
        """
        Инициализация API клиента.
        :param token: Bearer-токен True API
        :param apikey: API Key Национального каталога
        :param sandbox: использовать тестовую среду
        """
        self.token = token or os.getenv("TRUE_API_TOKEN")
        self.apikey = apikey or os.getenv("API_KEY")
        self.sandbox = sandbox

        if not self.token and not self.apikey:
            raise ValueError("Не найден ни token, ни apikey. "
                             "Установите переменные TRUE_API_TOKEN или API_KEY.")

        # Правильный хост API
        self.base_url = (
            "https://api.nk.sandbox.crptech.ru"
            if sandbox
            else "https://xn--80aqu.xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai"
        )

    # ---------------------------
    # Метод 1: Получить карточку по GTIN (v3/product)
    # ---------------------------
    def get_set_by_gtin(self, gtin: str):
        url = f"{self.base_url}/v3/product"
        headers = {"Accept": "application/json", "Content-Type": "application/json"}
        params = {"gtin": gtin, "format": "json"}

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.apikey:
            params["apikey"] = self.apikey

        logger.info(f"GET {url} (GTIN: {gtin})")
        response = requests.get(url, headers=headers, params=params, timeout=30)
        logger.info(f"Status: {response.status_code}")

        if response.status_code != 200:
            logger.error(f"Ошибка API: {response.status_code}")
            return None

        try:
            return response.json()
        except Exception as e:
            logger.error(f"Ошибка декодирования JSON: {e}")
            return None

    # ---------------------------
    # Метод 1: Получить карточку с расширенной информацией по GTIN (v3/feed-product)
    # ---------------------------
    def feedProduct(self, gtin: str):
        url = f"{self.base_url}/v3/feed-product"
        headers = {"Accept": "application/json", "Content-Type": "application/json"}
        params = {"gtin": gtin, "format": "json"}

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.apikey:
            params["apikey"] = self.apikey

        logger.info(f"GET {url} (GTIN: {gtin})")
        response = requests.get(url, headers=headers, params=params, timeout=30)
        logger.info(f"Status: {response.status_code}")

        if response.status_code != 200:
            logger.error(f"Ошибка API: {response.status_code}")
            return None

        try:
            return response.json()
        except Exception as e:
            logger.error(f"Ошибка декодирования JSON: {e}")
            return None

    # ---------------------------
    # Метод 2: Получить разрешительный документ (v4/rd-info-by-gtin)
    # ---------------------------
    def get_permit_document_by_gtin(self, gtin: str, inn: str):
        url = f"{self.base_url}/v4/rd-info-by-gtin"

        headers = {"Accept": "application/json", "Content-Type": "application/json"}
        params = {"format": "json"}
        payload = {"gtin": gtin, "inn": inn}

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.apikey:
            params["apikey"] = self.apikey

        logger.info(f"POST {url} (GTIN: {gtin})")
        response = requests.post(url, headers=headers, params=params, json=payload, timeout=30)
        logger.info(f"Status: {response.status_code}")

        if response.status_code != 200:
            logger.error(f"Ошибка API: {response.status_code}")
            return []

        try:
            data = response.json()
        except Exception as e:
            logger.error(f"Ошибка декодирования JSON: {e}")
            return []

        documents = []
        try:
            result = data.get("result", {})
            docs = result.get("documents", [])
            for d in docs:
                number = d.get("number")
                from_date = d.get("from_date")
                to_date = d.get("to_date")
                applicant = d.get("applicant")
                manufacturer = d.get("manufacturer")

                days_left = None
                if to_date:
                    try:
                        dt_to = datetime.strptime(to_date, "%Y-%m-%d")
                        days_left = (dt_to - datetime.now()).days
                    except ValueError:
                        logger.warning(f"Неверный формат даты: {to_date}")

                documents.append({
                    "number": number,
                    "from_date": from_date,
                    "to_date": to_date,
                    "days_left": days_left,
                    "applicant": applicant,
                    "manufacturer": manufacturer,
                })
        except Exception as e:
            logger.warning(f"Ошибка разбора структуры документа: {e}")

        return documents

    # ---------------------------
    # Метод 3: Получить список доступных GTIN для субаккаунта (v3/linked-gtins)
    # ---------------------------
    def get_linked_gtins(self, inn: str = None, gtin: str = None, limit: int = None, offset: int = None):
        """
        Получить список компаний и кодов товаров, по которым предоставлен доступ субаккаунту.
        
        :param inn: ИНН владельца товара (опционально)
        :param gtin: Код товара для проверки доступности (опционально)
        :param limit: Количество записей в ответе (макс. 10000)
        :param offset: Смещение относительно начала выдачи
        :return: Список доступных GTIN или None в случае ошибки
        """
        url = f"{self.base_url}/v3/linked-gtins"
        headers = {"Accept": "application/json", "Content-Type": "application/json"}
        params = {"format": "json"}

        # Добавляем опциональные параметры
        if inn:
            params["inn"] = inn
        if gtin:
            params["gtin"] = gtin
        if limit:
            params["limit"] = limit
        if offset:
            params["offset"] = offset

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.apikey:
            params["apikey"] = self.apikey

        logger.info(f"GET {url} (INN: {inn}, GTIN: {gtin}, limit: {limit}, offset: {offset})")
        
        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)
            logger.info(f"Status: {response.status_code}")

            if response.status_code != 200:
                logger.error(f"Ошибка API: {response.status_code}")
                return None

            data = response.json()
            
            # Обрабатываем результат
            result = data.get("result", {})
            linked_gtins = result.get("linked_gtins", [])
            errors = result.get("errors", [])
            
            # Логируем ошибки, если есть
            if errors:
                for error in errors:
                    logger.warning(f"Ошибка в ответе: {error.get('message')} (код: {error.get('code')})")
            
            return linked_gtins
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка сети: {e}")
            return None
        except Exception as e:
            logger.error(f"Ошибка обработки ответа: {e}")
            return None

    # ---------------------------
    # Метод 4: Получить все доступные GTIN с постраничной выгрузкой
    # ---------------------------
    def get_all_linked_gtins(self, inn: str = None, page_size: int = 1000):
        """
        Получить все доступные GTIN для субаккаунта с постраничной выгрузкой.
        
        :param inn: ИНН владельца товара (опционально)
        :param page_size: Размер страницы (макс. 10000)
        :return: Список всех доступных GTIN или None в случае ошибки
        """
        all_linked_gtins = []
        offset = 0
        
        logger.info(f"Начало постраничной выгрузки доступных GTIN (page_size: {page_size})")
        
        while True:
            logger.info(f"Запрос страницы с offset: {offset}")
            
            page_result = self.get_linked_gtins(
                inn=inn,
                limit=page_size,
                offset=offset
            )
            
            if page_result is None:
                logger.error("Ошибка при получении страницы, прерывание выгрузки")
                return None
                
            if not page_result:
                logger.info("Получена пустая страница, завершение выгрузки")
                break
                
            # Добавляем результаты текущей страницы
            all_linked_gtins.extend(page_result)
            logger.info(f"Получено {len(page_result)} GTIN на текущей странице, всего: {len(all_linked_gtins)}")
            
            # Проверяем, есть ли еще данные
            if len(page_result) < page_size:
                logger.info("Получено меньше запрошенного количества, завершение выгрузки")
                break
                
            # Увеличиваем offset для следующей страницы
            offset += page_size
            
            # Пауза между запросами для соблюдения лимитов API
            logger.info("Пауза 1 секунда перед следующим запросом...")
            import time
            time.sleep(1)
        
        logger.info(f"Постраничная выгрузка завершена. Всего получено GTIN: {len(all_linked_gtins)}")
        return all_linked_gtins


# ---------------------------
# Чтение GTIN из CSV или Excel
# ---------------------------
def load_gtin_list(filepath: str):
    """
    Загружает список GTIN из Excel (.xlsx) или CSV файла.
    Ожидается, что в файле есть колонка с названием 'gtin' (регистр не важен).
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Файл не найден: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()
    gtins = []

    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = [h.lower().strip() for h in reader.fieldnames]
            gtin_col = None
            for h in headers:
                if h in ("gtin", "код", "код_товара"):
                    gtin_col = h
                    break
            if not gtin_col:
                raise ValueError("Не найдена колонка 'gtin' или 'код_товара' в CSV.")

            for row in reader:
                val = row.get(gtin_col)
                if val:
                    gtins.append(val.strip())

    elif ext in (".xlsx", ".xls"):
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        gtin_col_idx = None
        for i, h in enumerate(headers):
            if h in ("gtin", "код", "код_товара"):
                gtin_col_idx = i
                break
        if gtin_col_idx is None:
            raise ValueError("Не найдена колонка 'gtin' или 'код_товара' в Excel.")

        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[gtin_col_idx]
            if val:
                gtins.append(str(val).strip())
    else:
        raise ValueError("Поддерживаются только файлы .csv и .xlsx")

    logger.info(f"Загружено {len(gtins)} GTIN из {filepath}")
    return gtins


# ---------------------------
# Сохранение результатов linked-gtins в CSV
# ---------------------------
def save_linked_gtins_to_csv(linked_gtins: list, output_file: str = None):
    """
    Сохраняет список доступных GTIN в CSV файл.
    """
    if not linked_gtins:
        logger.warning("Нет данных для сохранения")
        return None
        
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"linked_gtins_{timestamp}.csv"
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['gtin', 'producer_inn', 'producer_name']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in linked_gtins:
                writer.writerow({
                    'gtin': item.get('gtin', ''),
                    'producer_inn': item.get('producer_inn', ''),
                    'producer_name': item.get('producer_name', '')
                })
        
        logger.info(f"Результаты сохранены в файл: {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"Ошибка при сохранении в CSV: {e}")
        return None


# ---------------------------
# Функция для логирования статусов карточки
# ---------------------------
def log_product_status(item: dict, gtin: str):
    """
    Дополнительное логирование статусов карточки товара
    """
    try:
        # Логирование good_mark_flag
        good_mark_flag = item.get("good_mark_flag")
        if good_mark_flag is not None:
            if good_mark_flag:
                logger.info(f"GTIN {gtin}: Готов к заказу КМ")
            else:
                logger.warning(f"GTIN {gtin}: Не готов к заказу КМ")
        
        # Логирование good_turn_flag
        good_turn_flag = item.get("good_turn_flag")
        if good_turn_flag is not None:
            if good_turn_flag:
                logger.info(f"GTIN {gtin}: Готов к вводу в оборот")
            else:
                logger.warning(f"GTIN {gtin}: Не готов к вводу в оборот")
        
        # Логирование good_detailed_status
        good_detailed_status = item.get("good_detailed_status", [])
        if good_detailed_status:
            status_mapping = {
                "draft": ("черновик", logging.WARNING),
                "moderation": ("на модерации", logging.ERROR),
                "errors": ("требует изменений", logging.WARNING),
                "notsigned": ("ожидает подписания", logging.ERROR),
                "published": ("опубликована", logging.INFO)
            }
            
            for status in good_detailed_status:
                if status in status_mapping:
                    status_text, log_level = status_mapping[status]
                    if log_level == logging.INFO:
                        logger.info(f"GTIN {gtin}: Статус - {status_text}")
                    elif log_level == logging.WARNING:
                        logger.warning(f"GTIN {gtin}: Статус - {status_text}")
                    elif log_level == logging.ERROR:
                        logger.error(f"GTIN {gtin}: Статус - {status_text}")
                else:
                    logger.warning(f"GTIN {gtin}: Неизвестный статус - {status}")
        
        # Логирование good_status
        good_status = item.get("good_status", [])
        if good_status:
            if good_status in ["published"]:
                logger.info(f"GTIN {gtin}: Карточка опубликована")
            else:
                logger.error(f"GTIN {gtin}: статус - {good_status}")
        # Логирование flags_updated_date
        flags_updated_date = item.get("flags_updated_date")
        if flags_updated_date:
            logger.info(f"GTIN {gtin}: Дата обновления флагов - {flags_updated_date}")
            
    except Exception as e:
        logger.error(f"Ошибка при логировании статусов для GTIN {gtin}: {e}")


# ---------------------------
# CLI: запуск проверки
# ---------------------------
if __name__ == "__main__":
    # Определяем имя файла лога по умолчанию
    default_log_file = os.path.splitext(__file__)[0] + ".log"
    
    parser = argparse.ArgumentParser(
        description="Получение данных из Национального Каталога Честный Знак через True API"
    )
    parser.add_argument("--gtin", help="GTIN товара или набора")
    parser.add_argument("--file", help="Excel или CSV файл со списком GTIN")
    parser.add_argument("--inn", required=False, help="ИНН компании для проверки документа")
    parser.add_argument("--sandbox", action="store_true", help="использовать тестовую среду")
    parser.add_argument("--linked-gtins", action="store_true", help="получить ВЕСЬ список доступных GTIN для субаккаунта (постранично)")
    parser.add_argument("--linked-inn", help="ИНН владельца товара для фильтрации linked-gtins")
    parser.add_argument("--linked-limit", type=int, default=1000, help="Размер страницы для постраничной выгрузки linked-gtins (макс. 10000)")
    parser.add_argument("--linked-output", help="Файл для сохранения результатов linked-gtins (по умолчанию: linked_gtins_<timestamp>.csv)")
    parser.add_argument("--log-file", help=f"Файл для сохранения логов (по умолчанию: {default_log_file})")

    args = parser.parse_args()

    # Настраиваем логирование
    log_file = args.log_file if args.log_file else default_log_file
    setup_logging(log_file)

    try:
        nk = NK(sandbox=args.sandbox)

        # Обработка запроса linked-gtins с постраничной выгрузкой
        if args.linked_gtins:
            logger.info("Начало постраничной выгрузки ВСЕХ доступных GTIN для субаккаунта...")
            
            all_linked_gtins = nk.get_all_linked_gtins(
                inn=args.linked_inn,
                page_size=args.linked_limit
            )
            
            if all_linked_gtins is None:
                logger.error("Не удалось выгрузить список доступных GTIN")
                sys.exit(1)
                
            if not all_linked_gtins:
                logger.info("Нет доступных GTIN для данного субаккаунта")
            else:
                logger.info(f"Всего найдено {len(all_linked_gtins)} доступных GTIN:")
                
                # Сохраняем результаты в CSV
                output_file = save_linked_gtins_to_csv(all_linked_gtins, args.linked_output)
                
                # Выводим краткую статистику
                producers = {}
                for item in all_linked_gtins:
                    producer_inn = item.get('producer_inn')
                    if producer_inn not in producers:
                        producers[producer_inn] = {
                            'name': item.get('producer_name'),
                            'count': 0
                        }
                    producers[producer_inn]['count'] += 1
                
                logger.info("Статистика по производителям:")
                for inn, info in producers.items():
                    logger.info(f"  - {info['name']} (ИНН: {inn}): {info['count']} GTIN")
                
                if output_file:
                    logger.info(f"Полные результаты сохранены в: {output_file}")
                    
            sys.exit(0)

        # Остальная логика обработки GTIN
        gtin_list = []
        if args.file:
            gtin_list = load_gtin_list(args.file)
        elif args.gtin:
            gtin_list = [args.gtin]
        else:
            logger.error("Не указан GTIN или файл (--gtin или --file).")
            sys.exit(1)

        for gtin in gtin_list:
            logger.info("=" * 80)
            logger.info(f"Проверка GTIN: {gtin}")
            product = None
            try:
                product = nk.feedProduct(gtin=gtin)
            except Exception as e:
                logger.info(f'feed-product fo gtin{gtin}: {e}')
            if not product:
                product = nk.get_set_by_gtin(gtin)
            if product and product.get("result"):
                item = product["result"][0]
                logger.info(f"Наименование: {item.get('good_name')}")
                
                # ДОПОЛНИТЕЛЬНОЕ ЛОГИРОВАНИЕ СТАТУСОВ КАРТОЧКИ
                log_product_status(item, gtin)
                
                if item.get("is_set"):
                    logger.info("Это набор. Состав:")
                    for s in item.get("set_gtins"):
                        logger.info(f" - GTIN {s['gtin']} x{s['quantity']}")
            else:
                logger.warning("Карточка не найдена.")

            if args.inn:
                docs = nk.get_permit_document_by_gtin(gtin, args.inn)
                if not docs:
                    logger.warning(f"gtin:{gtin} Разрешительные документы не найдены.")
                else:
                    for d in docs:
                        logger.info(f"gtin:{gtin} Документ № {d['number']} от {d['from_date']} до {d['to_date']}")
                        if d["days_left"] is not None:
                            if d['days_left'] < 30:
                                logger.error(f"  Осталось {d['days_left']} дней до окончания")
                            else:
                                logger.info(f"  Осталось {d['days_left']} дней до окончания")
                        logger.info(f"  Заявитель: {d['applicant']}")
                        logger.info(f"  Изготовитель: {d['manufacturer']}")

    except Exception as e:
        logger.exception(f"Критическая ошибка: {e}")
        sys.exit(1)