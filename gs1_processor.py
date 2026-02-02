import csv
import json
import re
import glob
import logging
import argparse
import os
from pathlib import Path

# Попытка импорта openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger("gs1_processor")
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)

def get_gs1_prefix(gtin):
    """Выделяет префикс GS1 (стандартный 7-значный или специфический 9-значный)."""
    if not gtin:
        return None
    gtin_str = str(gtin).strip().split('.')[0]
    if not gtin_str.isdigit() or len(gtin_str) < 12:
        return None

    # Обработка специфических 9-значных префиксов
    if gtin_str.startswith("467001792"):
        return gtin_str[:9]

    return gtin_str[:7]

def extract_inn_from_filename(filename):
    """Извлекает ИНН из имени файла."""
    match = re.search(r'(\d{10,12})', filename)
    return match.group(1) if match else "unknown"

def get_inn_by_gtin(gtin, db_path='gs1prefix_inn_db.json'):
    """Выдает ИНН из базы JSON по номеру GTIN."""
    if not os.path.exists(db_path):
        return None
    try:
        with open(db_path, 'r', encoding='utf-8') as f:
            db = json.load(f)
        prefix = get_gs1_prefix(gtin)
        return db.get(prefix)
    except Exception as e:
        logger.error(f"Ошибка при поиске в базе: {e}")
        return None

def parse_csv(file_path, inn, prefix_map):
    try:
        with open(file_path, mode='r', encoding='utf-8', errors='ignore') as f:
            first_line = f.readline()
            if not first_line:
                return
            # Простейшее определение разделителя
            delimiter = ';' if ';' in first_line else ','

            # Читаем заголовки и приводим к нижнему регистру
            headers = [h.strip().lower() for h in first_line.split(delimiter)]

            reader = csv.DictReader(f, fieldnames=headers, delimiter=delimiter)
            for row in reader:
                prefix = get_gs1_prefix(row.get('gtin'))
                if prefix:
                    prefix_map[prefix] = inn
    except Exception as e:
        logger.error(f"Ошибка в CSV {file_path}: {e}")

def parse_xlsx(file_path, inn, prefix_map):
    if not HAS_OPENPYXL: return
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(max_row=1))]
        if 'gtin' in headers:
            gtin_idx = headers.index('gtin')
            for row in sheet.iter_rows(min_row=2, values_only=True):
                prefix = get_gs1_prefix(row[gtin_idx])
                if prefix: prefix_map[prefix] = inn
    except Exception as e:
        logger.error(f"Ошибка в XLSX {file_path}: {e}")

def build_prefix_inn_dict(masks):
    prefix_map = {}
    for mask in masks:
        for file_path in glob.glob(mask):
            filename = Path(file_path).name
            inn = extract_inn_from_filename(filename)
            if inn == "unknown": continue
            if filename.lower().endswith('.xlsx'):
                parse_xlsx(file_path, inn, prefix_map)
            else:
                parse_csv(file_path, inn, prefix_map)
    return prefix_map

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--masks', nargs='+', default=["owned_gtins*.csv", "linked_gtins*.csv", "owned_gtins*.xlsx"])
    parser.add_argument('--output', default='gs1prefix_inn_db.json')
    args = parser.parse_args()
    db = build_prefix_inn_dict(args.masks)
    if db:
        with open(args.output, 'w', encoding='utf-8') as jf:
            json.dump(db, jf, ensure_ascii=False, indent=4)
        logger.info(f"База сохранена: {args.output}")

if __name__ == "__main__":
    main()